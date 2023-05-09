import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import date


#startLine()
#returns the right most column in the excel sheet which is blank
#this is where we will start inputting new tickers found for the day
#PARAMETERS: currentPage -> the excel sheet we want to find the starting point of
#RETURNS: NONE
def startLine(currentPage):
    rowNumber = 1
    while currentPage.cell(row = rowNumber, column = 1).value != None:
        rowNumber += 1
    return rowNumber


#writeToExcel()
#using data retrieved by finvizData() will write to excel sheet
#PARAMETERS: data -> data scraped from Finviz, excelPage -> which page we want to write to
#RETURNS: NONE
def writeToExcel(data, excelPage):
    excelFile = load_workbook("D:/Programming/Repositories/finvizTracker/data.xlsx")
    currentPage = excelFile[excelPage]
 
    startingLine = startLine(currentPage)
    index = 1


    
    for ticker in data[0]: #for each row we have in our data aka each ticker
        currentPage.cell(row = startingLine, column = 1).value = data[0][index] #ticker
        currentPage.cell(row = startingLine, column = 2).value = data[1][index] #sector
        currentPage.cell(row = startingLine, column = 3).value = date.today() #date
        currentPage.cell(row = startingLine, column = 4).value = data[3][index] #cap
        currentPage.cell(row = startingLine, column = 5).value = float(data[4][index]) #price

        currentPage.cell(row = startingLine, column = 6).value = data[5][index] #change
        if float(data[5][index].replace("%","")) < 0: #check if change is positive or negative
            currentPage.cell(row = startingLine, column = 6).font = Font(color="c90000") #red
        else: 
            currentPage.cell(row = startingLine, column = 6).font = Font(color="26b013")#green

        currentPage.cell(row = startingLine, column = 6).value = data[5][index] #change
        currentPage.cell(row = startingLine, column = 7).value = float(data[6][index]) #volume
        index += 1
        startingLine += 1
    excelFile.save("D:/Programming/Repositories/finvizTracker/data.xlsx")

#parseData()
#retrieves finviz screener data by utilizing pandas to read html tables on finviz.com
#after gaining the data we want, we put it in an array to pass to writeToExcel()
#PARAMETERS: url -> Finviz page we want to scrape, excelPage -> which page we want to write to
#RETURNS: NONE
def parseData(url, excelPage):
    header = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
    finvizPage = requests.get(url, headers=header).text  
    tables = pd.read_html(finvizPage)
    if len(tables) == 21: #if our finviz settings return no tickers, there will only be 20 tables on the website
        table = tables[-2] #this is the table we want from the many tables pandas found
        names = table.iloc[1:, 1] #I think _: is better than 1: even if do the same thing
        sectors = table.iloc[1:, 3]
        marketCaps = table.iloc[1:,6]
        prices = table.iloc[1:,8]
        changes = table.iloc[1:,9]
        volumes = table.iloc[1:,10] #[row selection, column selection] so 1:,10 means every row, but only column 10

       
        
        data = [names, sectors, date.today(), marketCaps, prices, changes, volumes]
   
        writeToExcel(data, excelPage)
    else:
        print("NO TICKERS FOOUND ON " + url)



