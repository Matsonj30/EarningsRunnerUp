from openpyxl import load_workbook
import pyodbc
from yahoo_fin.stock_info import *
from datetime import date
from main import parseData


def updateSheets():

    excelFile = load_workbook("D:/Programming/Repositories/finvizTracker/data.xlsx")

    for worksheet in excelFile.worksheets:
        index = 2
  
        currentPage = excelFile[worksheet.title]
        while(currentPage.cell(row=index, column=1).value != None):
            if(currentPage.cell(row=index, column=21).value == None):
                for i in range(14):
                    if(currentPage.cell(row=index, column= i + 8).value == None):
                        currentPage.cell(row=index, column=i+ 8 ).value = round(get_live_price( currentPage.cell(row=index, column=1).value),2)
                        break
            index += 1
    excelFile.save("D:/Programming/Repositories/finvizTracker/data.xlsx")

updateSheets()
parseData('https://finviz.com/screener.ashx?v=111&f=cap_microover,geo_usa,sh_float_u100,sh_price_o1,sh_relvol_o5,ta_sma200_pb&ft=4&o=volume', 'IrregularVolume')
parseData('https://finviz.com/screener.ashx?v=111&f=cap_microover,geo_canada,sh_float_u100,sh_relvol_o5,ta_sma200_pb&ft=4&o=volume', 'IrregularVolume')