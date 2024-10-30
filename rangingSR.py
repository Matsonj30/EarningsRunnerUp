import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import date


#startLine()
#returns the right most column in the excel sheet which is blank
#this is where we will start inputting new tickers found for the day
#PARAMETERS: currentPage -> the excel sheet we want to find the starting point of
#RETURNS: NONE
def startLine(currentPage):
    rowNumber = 1
    while currentPage.cell(row = rowNumber, column = 1).value != None:
        rowNumber += 1
    return rowNumber


#def findAllTickers
#appends all tickers in the excel sheet to a dictionary, allowing us to easily filter new tickers on the finviz screener
def findAllTickers(excelPage):
    excelFile = load_workbook("D:/Programming/Repositories/finvizTracker/RangingSR.xlsx")
    currentPage = excelFile[excelPage]
    rowNumber = 2
    tickerDictionary = {}
    while currentPage.cell(row = rowNumber, column = 1).value != None:
        tickerDictionary[currentPage.cell(row = rowNumber, column = 1).value] = None #append to dictionary for immediate finding
        rowNumber += 1
    print(tickerDictionary)
    return tickerDictionary
    

#writeToExcel()
#using data retrieved by finvizData() will write to excel sheet
#PARAMETERS: data -> data scraped from Finviz, excelPage -> which page we want to write to
#RETURNS: NONE
def writeToExcel(data, excelPage, alreadyFoundTickers):
    excelFile = load_workbook("D:/Programming/Repositories/finvizTracker/RangingSR.xlsx")
    currentPage = excelFile[excelPage]
 
    startingLine = startLine(currentPage)
    index = 0

    for ticker in data[0]: #for each row we have in our data aka each ticker
        #only add new findings not in the sheet already, this screener has little variance
        if(ticker not in alreadyFoundTickers):
            currentPage.cell(row = startingLine, column = 1).value = data[0][index] #ticker
            #currentPage.cell(row = startingLine, column = 2).value = data[1][index] #sector
            currentPage.cell(row = startingLine, column = 3).value = date.today() #date
            #currentPage.cell(row = startingLine, column = 4).value = data[3][index] #cap
            #currentPage.cell(row = startingLine, column = 5).value = float(data[4][index]) #price
            #currentPage.cell(row = startingLine, column = 6).value = data[5][index] #change
            #if float(data[5][index].replace("%","")) < 0: #check if change is positive or negative
            #    currentPage.cell(row = startingLine, column = 6).font = Font(color="c90000") #red
            #else: 
            #    currentPage.cell(row = startingLine, column = 6).font = Font(color="26b013")#green

            #currentPage.cell(row = startingLine, column = 6).value = data[5][index] #change
            #currentPage.cell(row = startingLine, column = 7).value = float(data[6][index]) #volume
            index += 1
            startingLine += 1
    excelFile.save("D:/Programming/Repositories/finvizTracker/RangingSR.xlsx")



#parseData()
#retrieves finviz screener data by utilizing pandas to read html tables on finviz.com
#after gaining the data we want, we put it in an array to pass to writeToExcel()
#PARAMETERS: url -> Finviz page we want to scrape, excelPage -> which page we want to write to
#RETURNS: NONE
def parseData(excelPage):
    header = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
    urls = ["https://finviz.com/screener.ashx?v=111&p=d&f=cap_largeover,geo_usa,sh_avgvol_o1000,sh_opt_option,sh_price_o2,ta_averagetruerange_u0.5,ta_pattern_horizontal&ft=4&ta=0", "https://finviz.com/screener.ashx?v=111&p=d&f=cap_largeover,geo_usa,sh_avgvol_o1000,sh_opt_option,sh_price_o2,ta_averagetruerange_u0.5,ta_pattern_horizontal&ft=4&ta=0&r=21", "https://finviz.com/screener.ashx?v=111&p=d&f=cap_largeover,geo_usa,sh_avgvol_o1000,sh_opt_option,sh_price_o2,ta_averagetruerange_u0.5,ta_pattern_horizontal&ft=4&ta=0&r=41"]
    alreadyFoundTickers = findAllTickers(excelPage)

    for url in urls:
        finvizPage = requests.get(url, headers=header).text  
        tables = pd.read_html(finvizPage)
        if len(tables) == 28: #if our finviz settings return no tickers, there will only be 20 tables on the website
            table = tables[-2] #this is the table we want from the many tables pandas found

            names = table.iloc[:, 1] #I think _: is better than 1: even if do the same thing

            sectors = table.iloc[:, 3]
            marketCaps = table.iloc[:,6]
            prices = table.iloc[:,8]
            changes = table.iloc[:,9]
            volumes = table.iloc[:,10] #[row selection, column selection] so 1:,10 means every row, but only column 10

            
            
            data = [names, sectors, date.today(), marketCaps, prices, changes, volumes]
            
            writeToExcel(data, excelPage, alreadyFoundTickers)
        else:
            print("NO TICKERS FOUND ON " + url)



parseData("Ranging")